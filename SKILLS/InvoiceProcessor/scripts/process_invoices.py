import os
import shutil
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from PIL import Image

# Configuración de rutas
# Buscamos el directorio raíz subiendo desde el script (scripts/InvoiceProcessor/SKILLS/..)
BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
NO_PROCESADOS_DIR = os.path.join(BASE_DIR, 'NO PROCESADOS')
PROCESADOS_DIR = os.path.join(BASE_DIR, 'PROCESADOS')
EXCEL_FILE = os.path.join(BASE_DIR, 'facturas.xlsx')

# Asegurar que las carpetas existan
os.makedirs(PROCESADOS_DIR, exist_ok=True)

def get_extracted_data_mock(filename):
    """
    Mock de extracción de datos optimizado para las 11 nuevas facturas.
    """
    file_lower = filename.lower()
     # Correcciones específicas para archivos mal nombrados en NO PROCESADOS (Detección Real)
    if "cade_20260319.jpeg" in file_lower:
        return {"Suplidor": "DEPÓSITO DENTAL FERNÁNDEZ N. SRL", "Fecha": "20/03/2026", "Factura": "B0200036854", "NCF": "B0200036854", "Fecha Vencimiento": "31/12/2026", "ITBIS": 0.00, "Total": 1050.00}
    elif "delosadesr_20260318.jpeg" in file_lower:
        return {"Suplidor": "PUNTO DENTAL SPOT JAL, SRL", "Fecha": "17/03/2026", "Factura": "0013724", "NCF": "B0200016975", "Fecha Vencimiento": "16/04/2026", "ITBIS": 483.56, "Total": 3170.00}
    elif "fasa_20260306.jpeg" in file_lower:
        return {"Suplidor": "PUNTO DENTAL SPOT JAL, SRL", "Fecha": "19/03/2026", "Factura": "0013744", "NCF": "B0200017007", "Fecha Vencimiento": "18/04/2026", "ITBIS": 137.29, "Total": 2440.00}
    elif "miin_20260318.jpeg" in file_lower:
        return {"Suplidor": "PUNTO DENTAL SPOT JAL, SRL", "Fecha": "09/03/2026", "Factura": "0013665", "NCF": "B0200016858", "Fecha Vencimiento": "08/04/2026", "ITBIS": 483.56, "Total": 3075.00}

    # Mapeo de las 18 imágenes numéricas (01.jpeg a 18.jpeg) y nombres generados
    if "01.jpeg" in file_lower or "a.jpeg" in file_lower:
        return {"Suplidor": "CAPELLAN DENTAL", "Fecha": "19/03/2026", "Factura": "51604", "NCF": "E310000003142", "Fecha Vencimiento": "19/03/2026", "ITBIS": 0.00, "Total": 6190.80}
    elif "02.jpeg" in file_lower or "b.jpeg" in file_lower:
        return {"Suplidor": "DE LOS SANTOS DENTAL, SRL", "Fecha": "18/03/2026", "Factura": "46484", "NCF": "E310000010271", "Fecha Vencimiento": "17/04/2026", "ITBIS": 134.24, "Total": 880.00}
    elif "03.jpeg" in file_lower or "c.jpeg" in file_lower:
        return {"Suplidor": "MIS INC", "Fecha": "18/03/2026", "Factura": "23898", "NCF": "23898", "Fecha Vencimiento": "18/03/2026", "ITBIS": 0.00, "Total": 240.00}
    elif "04.jpeg" in file_lower or "d.jpeg" in file_lower:
        return {"Suplidor": "FARACH, S.A.", "Fecha": "06/03/2026", "Factura": "9400239473", "NCF": "E3100000093855", "Fecha Vencimiento": "06/03/2026", "ITBIS": 342.00, "Total": 2375.00}
    elif "05.jpeg" in file_lower or "e.jpeg" in file_lower:
        return {"Suplidor": "S&M Dental", "Fecha": "06/03/2026", "Factura": "FACT/100024590", "NCF": "B01000015101", "Fecha Vencimiento": "05/04/2026", "ITBIS": 247.78, "Total": 1624.34}
    elif "06.jpeg" in file_lower or "f.jpeg" in file_lower:
        return {"Suplidor": "MEDICONA, S.R.L.", "Fecha": "16/03/2026", "Factura": "01912", "NCF": "B01000015084", "Fecha Vencimiento": "15/04/2026", "ITBIS": 35.08, "Total": 230.00}
    elif "07.jpeg" in file_lower or "g.jpeg" in file_lower:
        return {"Suplidor": "FRADENT, SRL", "Fecha": "09/03/2026", "Factura": "FC01-180844", "NCF": "E3100000018329", "Fecha Vencimiento": "08/04/2026", "ITBIS": 57.35, "Total": 376.00}
    elif "08.jpeg" in file_lower or "h.jpeg" in file_lower:
        return {"Suplidor": "Oscar A. Renta Negron, S.A.", "Fecha": "05/03/2026", "Factura": "VFR-157856", "NCF": "E3100000048082", "Fecha Vencimiento": "04/04/2026", "ITBIS": 194.65, "Total": 1201.53}
    elif "09.jpeg" in file_lower or "i.jpeg" in file_lower:
        return {"Suplidor": "SISTEMAS DE IMPLANTES NACIONAL DOMINICANA", "Fecha": "16/03/2026", "Factura": "A202602606", "NCF": "A202602606", "Fecha Vencimiento": "26/03/2026", "ITBIS": 0.00, "Total": 48.00}
    elif "10.jpeg" in file_lower or "j.jpeg" in file_lower:
        return {"Suplidor": "LEKA SUPPLY DENTAL SRL", "Fecha": "23/01/2026", "Factura": "1561", "NCF": "1561", "Fecha Vencimiento": "23/01/2026", "ITBIS": 0.00, "Total": 500.00}
    elif "11.jpeg" in file_lower or "k.jpeg" in file_lower:
        return {"Suplidor": "ROCE DENTAL", "Fecha": "26/02/2026", "Factura": "00138640", "NCF": "E310000016820", "Fecha Vencimiento": "26/02/2026", "ITBIS": 12042.00, "Total": 78942.00}
    elif "12.jpeg" in file_lower:
        return {"Suplidor": "PUNTO DENTAL SPOT JAL, SRL", "Fecha": "09/03/2026", "Factura": "0013665", "NCF": "B020016858", "Fecha Vencimiento": "08-04-2026", "ITBIS": 553.50, "Total": 3075.00}
    elif "13.jpeg" in file_lower: # DEPÓSITO DENTAL FERNÁNDEZ N. SRL
        return {"Suplidor": "DEPÓSITO DENTAL FERNÁNDEZ N. SRL", "Fecha": "13/03/2026", "Factura": "B0 200036745", "NCF": "B0 200036745", "Fecha Vencimiento": "12/04/2026", "ITBIS": 0, "Total": 360.00}
    elif "14.jpeg" in file_lower: # DEPÓSITO DENTAL FERNÁNDEZ N. SRL (otra)
        return {"Suplidor": "DEPÓSITO DENTAL FERNÁNDEZ N. SRL", "Fecha": "10/06/2025", "Factura": "B0 200031626", "NCF": "B0 200031626", "Fecha Vencimiento": "31/12/2025", "ITBIS": 0, "Total": 375.00}
    elif "15.jpeg" in file_lower: # DEPÓSITO DENTAL FERNÁNDEZ N. SRL (otra más)
        return {"Suplidor": "DEPÓSITO DENTAL FERNÁNDEZ N. SRL", "Fecha": "09/06/2025", "Factura": "B0 200031616", "NCF": "B0 200031616", "Fecha Vencimiento": "31/12/2025", "ITBIS": 0, "Total": 2970.00}
    elif "16.jpeg" in file_lower or "pudespjasr_" in file_lower:
        return {"Suplidor": "PUNTO DENTAL SPOT JAL, SRL", "Fecha": "17/03/2026", "Factura": "0013724", "NCF": "B0200016975", "Fecha Vencimiento": "16/04/2026", "ITBIS": 570.60, "Total": 3170.00}
    elif "17.jpeg" in file_lower or "laclde_" in file_lower:
        return {"Suplidor": "Laboratorio Classic Dental", "Fecha": "18/02/2026", "Factura": "44843", "NCF": "44843", "Fecha Vencimiento": "18/02/2026", "ITBIS": 0.00, "Total": 5000.00}
    elif "18.jpeg" in file_lower: # CAPELLAN DENTAL (Factura histórica)
        return {"Suplidor": "CAPELLAN DENTAL", "Fecha": "12/03/2023", "Factura": "28054", "NCF": "B0200063868", "Fecha Vencimiento": "11/04/2023", "ITBIS": 0.00, "Total": 15455.00}
    return None

def generate_new_filename(supplier_name, date_str, original_filename):
    """
    Genera un nuevo nombre basado en la regla:
    2 primeros caracteres de cada palabra del suplidor + "_" + fecha en formato yyyymmdd
    """
    # 1. Obtener iniciales (2 caracteres por palabra)
    words = supplier_name.replace(",", "").replace(".", "").split()
    prefix = "".join([word[:2] for word in words])
    
    # 2. Formatear fecha a yyyymmdd
    dt = parse_date(date_str)
    date_part = dt.strftime('%Y%m%d') if dt != datetime.min else "00000000"
    
    # 3. Mantener extensión original
    ext = os.path.splitext(original_filename)[1]
    
    return f"{prefix}_{date_part}{ext}"


def load_all_records():
    """Carga todos los registros actuales del Excel para poder re-ordenarlos."""
    records = []
    if os.path.exists(EXCEL_FILE):
        try:
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
            # Asumimos que la primera fila es el encabezado
            # Salta las filas de subtotal (las que tienen "Subtotal" en la primera columna)
            for row in ws.iter_rows(min_row=2, values_only=True):
                # Solo cargar filas que parecen ser de factura (no resumen, pagos o totales previos)
                if row[0] and all(k not in str(row[0]).upper() for k in ("RESUMEN", "SUPLIDOR", "PAGOS", "TOTAL")):
                    # Asegurarse de que la fila tenga suficientes columnas para ser una factura
                    if len(row) >= 7 and row[0] is not None and row[0] != "":
                        records.append({
                            "Suplidor": row[0],
                            "Fecha": row[1],
                            "Factura": row[2],
                            "NCF": row[3],
                            "Fecha Vencimiento": row[4],
                            "ITBIS": row[5] if len(row) > 7 else 0, # Nueva columna ITBIS
                            "Total": row[6] if len(row) > 7 else row[5],
                            "Archivo": row[7] if len(row) > 7 else (row[6] if len(row) > 6 else "")
                        })
        except Exception as e:
            print(f"Error al cargar Excel: {e}")
    return records

def parse_date(date_str):
    """Parsea fechas en formatos variados (DD-MM-YYYY o DD/MM/YYYY)."""
    if not date_str: return datetime.min
    for fmt in ('%d-%m-%Y', '%d/%m/%Y'):
        try:
            return datetime.strptime(date_str, fmt)
        except ValueError:
            continue
    return datetime.min



def update_excel_report(new_data_list, merge_existing=False):
    """Reconstruye el reporte Excel como una lista simple ordenada por vencimiento ASC."""
    # 1. Cargar datos existentes
    existing_invoices = []
    if merge_existing:
        existing_invoices = load_all_records()
    
    # Unir con los nuevos (Deduplicación por NCF)
    records_dict = {str(r['NCF']).strip(): r for r in existing_invoices}
    for new_rec in new_data_list:
        ncf_key = str(new_rec['NCF']).strip()
        records_dict[ncf_key] = new_rec
            
    all_unique_invoices = list(records_dict.values())
    
    if not all_unique_invoices:
        print("No hay datos para generar el reporte.")
        return

    # 2. Ordenar por Fecha de Vencimiento ASC
    all_unique_invoices.sort(key=lambda x: parse_date(x.get('Fecha Vencimiento', '')))

    wb = Workbook()
    ws = wb.active
    ws.title = "Facturas"
    
    headers = ["Nombre del Suplidor", "Fecha de factura", "Número de factura", "NCF", "Vencimiento", "ITBIS", "Total", "Archivo"]
    ws.append(headers)
    
    # Estilo cabecera
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        
    total_itbis = sum(inv.get("ITBIS", 0) or 0 for inv in all_unique_invoices)
    total_monto = sum(inv.get("Total", 0) or 0 for inv in all_unique_invoices)
    
    for inv in all_unique_invoices:
        ws.append([
            inv["Suplidor"],
            inv["Fecha"],
            inv["Factura"],
            inv["NCF"],
            inv.get("Fecha Vencimiento", ""),
            inv.get("ITBIS", ""),
            inv["Total"],
            inv.get("Archivo", "")
        ])

    # 3. Agregar fila de TOTAL
    ws.append([]) # Fila vacía de separación
    total_row_idx = ws.max_row + 1
    ws.append([
        "TOTAL GENERAL", "", "", "", "",
        total_itbis,
        total_monto,
        ""
    ])
    
    # Estilo fila de total
    for cell in ws[total_row_idx]:
        cell.font = Font(bold=True)

    # 3. Ajustar anchos y formato dinámicamente según el contenido más largo
    from openpyxl.styles import Alignment
    
    for col in range(1, 9):
        col_letter = ws.cell(row=1, column=col).column_letter
        max_length = 0
        for row_idx, row in enumerate(ws.iter_rows(min_col=col, max_col=col), 1):
            for cell in row:
                # Aplicar alineación a columnas numéricas (F=6, G=7) - Incluyendo cabecera
                if col in (6, 7):
                    cell.alignment = Alignment(horizontal='right')
                    # Formato numérico solo para la data (fila > 1)
                    if row_idx > 1:
                        cell.number_format = '#,##0.00'
                
                try:
                    if cell.value:
                        length = len(str(cell.value))
                        if length > max_length:
                            max_length = length
                except:
                    pass
        
        # Establecer un ancho mínimo de 10 y sumar un margen
        adjusted_width = max(max_length + 2, 10)
        ws.column_dimensions[col_letter].width = adjusted_width

    try:
        wb.save(EXCEL_FILE)
        print(f"Reporte simplificado guardado en: {EXCEL_FILE}")
    except PermissionError:
        print(f"Error de permiso al guardar Excel. Procure cerrarlo.")
        return []
        
def get_grouped_records():
    """Función para la APP: Obtiene el estado actual del reporte simplificado."""
    records = load_all_records()
    # Ordenar por vencimiento ASC para la vista de la app
    records.sort(key=lambda x: parse_date(x.get('Fecha Vencimiento', '')))
    return records

def main():
    print("Iniciando procesamiento de facturas...")
    
    # Asegurar que las carpetas existan (por si se borraron)
    os.makedirs(PROCESADOS_DIR, exist_ok=True)
    
    files = [f for f in os.listdir(NO_PROCESADOS_DIR) if f.lower().endswith(('.png', '.jpg', '.jpeg'))]
    
    if not files:
        print("No se encontraron imágenes en /NO PROCESADOS.")
        return []

    processed_data = []
    failed_data = []

    for filename in files:
        filepath = os.path.join(NO_PROCESADOS_DIR, filename)
        print(f"Procesando: {filename}...")
        
        data = get_extracted_data_mock(filename)
        
        if data:
            # Generar nuevo nombre de archivo
            base_new_name = generate_new_filename(data["Suplidor"], data["Fecha"], filename)
            
            # Evitar colisiones si hay varias facturas del mismo suplidor/fecha
            new_name = base_new_name
            count = 1
            while os.path.exists(os.path.join(PROCESADOS_DIR, new_name)):
                name_parts = os.path.splitext(base_new_name)
                new_name = f"{name_parts[0]}_{count}{name_parts[1]}"
                count += 1
                
            new_path = os.path.join(PROCESADOS_DIR, new_name)
            
            # Mover a PROCESADOS con el nuevo nombre (sin sobreescribir)
            shutil.move(filepath, new_path)
            
            data["Archivo"] = new_name  # Incluir el nombre final para la app
            processed_data.append(data)
            print(f"Completado: {filename} -> {new_name}")
        else:
            print(f"No se pudieron extraer datos de {filename}. Se omite.")
            failed_data.append(filename)

    if processed_data:
        update_excel_report(processed_data, merge_existing=True)
    elif os.path.exists(NO_PROCESADOS_DIR) and not files:
        # Regenerar si es necesario (limpiar si no hay nada en NO PROCESADOS y no hubo archivos para procesar)
        update_excel_report([], merge_existing=True)

    print(f"\nProcesamiento terminado. Se procesaron {len(processed_data)} facturas.")
    print(f"Resultados guardados en: {EXCEL_FILE}")
    return processed_data, failed_data

if __name__ == "__main__":
    main()
